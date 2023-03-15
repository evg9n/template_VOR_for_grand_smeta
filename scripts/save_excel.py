from datetime import datetime
from json import dump, dumps
from os.path import join
from os import mkdir, path
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl import load_workbook
from openpyxl import worksheet
from typing import Union, Dict, List
from shutil import move
import logging


log = logging.getLogger()


def border(cell: worksheet) -> None:
    black = Side(border_style='thin')  # Установка стиял границ тонкие и черные
    cell.border = Border(top=black, bottom=black, left=black, right=black)  # Установка всех границ


def write_table(activ: worksheet, table: Union[List, Dict]) -> None:
    row = 18

    for chapter in table:
        row += 1

        list_positions = list(chapter.values())[0]

        if not list_positions:
            continue

        for col in range(1, 8):
            border(activ.cell(row, 1))  # Устанавливаем границы на ячейку

        activ.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        activ.cell(row, 1).value = list(chapter.keys())[0]
        activ.cell(row, 1).font = Font(bold=True)  # Установка жирного шрифта
        activ.cell(row, 1).alignment = Alignment(horizontal='left', vertical='center')  # Выравнивание лево центр

        for position in list_positions:
            row += 1
            activ.insert_rows(row)
            activ.cell(row, 1).value = position['number_pp']
            activ.cell(row, 2).value = position['number']
            activ.cell(row, 3).value = position['caption']
            activ.cell(row, 4).value = position['units']
            activ.cell(row, 5).value = position['result']
            activ.cell(row, 7).value = position['fx']

            for col in range(1, 8):
                border(activ.cell(row, col))
                if col == 3:
                    activ.cell(row, 3).alignment = Alignment(horizontal='left', vertical='top', wrapText=True)
                    # Выравниевание по лево и верх с переносом текста
                else:
                    activ.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    # Выравнивание по центру


def write_date(sheet: worksheet, row: int, col: int) -> None:

    months = {
        1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
        7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря',

    }

    today = datetime.today()
    year = today.year
    month = today.month
    day = today.day
    sheet.cell(row, col).value = f'"{day}" {months[month]} {year} г.'


def add_history(name: str) -> None:

    if not path.exists('history'):
        mkdir('history')

    move(join('download', f"{name}.xml"), join('history', f"{name}.xml"))


def save_excel(name: str, name_form: str, info_smeta: Union[List, Dict]) -> None:

    # print(dumps(info_smeta, indent=4, ensure_ascii=False))

    excel_file = join('patterns', name_form)
    a = load_workbook(excel_file)

    activ = a.active

    write_date(activ, 4, 1)
    write_date(activ, 4, 5)

    table = info_smeta.get('table')

    activ.cell(row=22, column=1).value = f"Составил:  _________________ /{info_smeta.get('составил') or ''}/"
    activ.cell(row=24, column=1).value = f"Проверил:  _________________ /{info_smeta.get('проверил') or ''}/"
    activ.cell(row=12, column=1).value = f"Вид объекта: {info_smeta.get('объект') or name}"
    # activ.cell().value = file_json.get()
    # activ.cell().value = file_json.get()
    # activ.cell().value = file_json.get()
    # activ.cell().value = file_json.get()

    write_table(activ, table)

    if not path.exists('result'):
        mkdir('result')
    new_file = join('result', f'{name}.xlsx')
    a.save(new_file)

    add_history(name)
